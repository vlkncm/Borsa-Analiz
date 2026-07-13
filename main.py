from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import time
import os
import sys

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from borsa_tarayici import teknik_analiz, rapor_yazdir
from backtest import backtest_toplu
from kap_modulu import kap_toplu_analiz
from mtf_grafik import coklu_zaman_dilimi_analizi, grafik_toplu_olustur
from olasilik_temettu import olasilik_toplu_ekle, temettu_toplu_tara
from faaliyet_raporu import faaliyet_toplu_analiz, faaliyet_dataframe
from piyasa_guncelleme import guncel_hisse_dosyasi
from v4_puanlama import v4_toplu_puanla

YASAL_UYARI_KISA = "Bu yazılım ve rapor yatırım tavsiyesi değildir; genel nitelikte algoritmik karar destek çıktısıdır. Kesin getiri garantisi vermez. Tüm yatırım kararları ve risk kullanıcıya aittir."

YASAL_UYARI_UZUN = [
    ["BORSA ANALİZ PRO MAX — YASAL UYARI"],
    ["Bu yazılım yatırım danışmanlığı, portföy yöneticiliği veya kişiye özel yatırım tavsiyesi hizmeti sunmaz."],
    ["Puanlar, sıralamalar, hedefler, stoplar ve olasılıklar genel nitelikte algoritmik karar destek çıktılarıdır."],
    ["Geçmiş performans ve backtest sonuçları gelecekteki performansın göstergesi veya garantisi değildir."],
    ["Veriler gecikebilir, eksik veya hatalı olabilir; işlem öncesinde resmi kaynaklardan doğrulanmalıdır."],
    ["Tüm yatırım kararları, finansal sonuçlar ve riskler kullanıcıya aittir."],
]

from pro_moduller import (
    temel_analiz_yfinance,
    haber_analizi_yfinance,
    makro_analiz_yfinance,
    broker_karar_motoru,
    portfoy_onerisi_uret
)



def kullanici_veri_klasoru():
    """Raporları Program Files yerine kullanıcının Belgeler klasörüne kaydeder."""
    documents = Path.home() / "Documents"
    if not documents.exists():
        documents = Path.home()
    base = documents / "Borsa Analiz Pro MAX"
    (base / "output").mkdir(parents=True, exist_ok=True)
    (base / "output" / "grafikler").mkdir(parents=True, exist_ok=True)
    (base / "logs").mkdir(parents=True, exist_ok=True)
    return base


def output_klasoru():
    return kullanici_veri_klasoru() / "output"

def resource_path(relative_path):
    """
    PyInstaller tek EXE içinde gömülü gelen dosyaları bulur.
    Normal çalıştırmada proje klasörünü kullanır.
    """
    try:
        base_path = Path(sys._MEIPASS)
    except Exception:
        base_path = Path(__file__).parent

    return base_path / relative_path

def dogrulanmis_hisse_dosyasi():
    documents = Path.home() / "Documents"
    if not documents.exists():
        documents = Path.home()
    path = documents / "Borsa Analiz Pro MAX" / "piyasa_verileri" / "bist_hisseleri_dogrulanmis.txt"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def hisseleri_txt_oku(dosya_adi="bist_hisseleri_613_aktif.txt"):
    """
    v4 öncelik sırası:
    1) Son başarılı taramada doğrulanmış gerçek fiyat verisi olan hisseler
    2) Paketle gelen doğrulanmış başlangıç listesi
    3) Güncel KAP listesi
    4) Eski gömülü liste
    """
    user_verified = dogrulanmis_hisse_dosyasi()
    if user_verified.exists():
        try:
            hisseler = [
                x.strip() for x in user_verified.read_text(encoding="utf-8").splitlines()
                if x.strip()
            ]
            if len(hisseler) >= 450:
                print(f"Doğrulanmış v4 hisse listesi kullanılıyor: {len(hisseler)} sembol")
                return sorted(set(hisseler))
        except Exception as exc:
            print(f"Doğrulanmış liste okunamadı: {exc}")

    packaged_verified = resource_path("bist_hisseleri_dogrulanmis.txt")
    if packaged_verified.exists():
        try:
            hisseler = [
                x.strip() for x in packaged_verified.read_text(encoding="utf-8").splitlines()
                if x.strip()
            ]
            if len(hisseler) >= 450:
                print(f"Paket doğrulanmış hisse listesi kullanılıyor: {len(hisseler)} sembol")
                return sorted(set(hisseler))
        except Exception:
            pass

    guncel_yol = guncel_hisse_dosyasi()
    if guncel_yol.exists():
        try:
            hisseler = [
                x.strip() for x in guncel_yol.read_text(encoding="utf-8").splitlines()
                if x.strip()
            ]
            if len(hisseler) >= 300:
                print(f"KAP güncel listesi kullanılıyor: {len(hisseler)} sembol")
                return sorted(set(hisseler))
        except Exception as exc:
            print(f"KAP listesi okunamadı: {exc}")

    dosya_yolu = resource_path(dosya_adi)
    hisseler = [
        x.strip() for x in dosya_yolu.read_text(encoding="utf-8").splitlines()
        if x.strip()
    ]
    print(f"Gömülü yedek liste kullanılıyor: {len(hisseler)} sembol")
    return sorted(set(hisseler))


def dogrulanmis_listeyi_kaydet(results):
    symbols = sorted({
        str(item.get("symbol", "")).strip()
        for item in results
        if str(item.get("symbol", "")).strip().endswith(".IS")
    })
    if len(symbols) < 450:
        print(f"Doğrulanmış liste kaydedilmedi; sayı yetersiz: {len(symbols)}")
        return
    hedef = dogrulanmis_hisse_dosyasi()
    tmp = hedef.with_suffix(".tmp")
    tmp.write_text("\n".join(symbols) + "\n", encoding="utf-8")
    tmp.replace(hedef)
    print(f"Doğrulanmış hisse listesi kaydedildi: {len(symbols)} sembol")


def hisse_tara(symbol):
    return teknik_analiz(symbol, "BIST GENEL")


def sonuclari_sirala(results):
    return sorted(
        results,
        key=lambda x: (
            x.get("v4_guven_puani", x.get("broker_skor", x.get("genel_skor", x.get("guven", 0)))),
            x.get("guven", 0),
            x.get("risk_getiri_1", 0)
        ),
        reverse=True
    )


def profesyonel_veri_ekle(results):
    print("\nMakro analiz yapılıyor...")
    makro = makro_analiz_yfinance()
    print(f"Makro: {makro.get('makro_etiket')} | {makro.get('makro_puan')}/100")

    # Performans için tüm 613 yerine en güçlü ilk adaylara temel/haber/KAP ekliyoruz.
    # Varsayılan 30; app.py üzerinden env ile değiştirilebilir.
    analiz_limiti = int(os.environ.get("PRO_ANALIZ_LIMIT", os.environ.get("KAP_ANALIZ_LIMIT", "30")))

    oncelikli = sorted(results, key=lambda x: x.get("guven", 0), reverse=True)
    zenginlestirilecek_liste = [x["symbol"] for x in oncelikli[:analiz_limiti]]
    zenginlestirilecek = set(zenginlestirilecek_liste)

    print(f"Pro analiz limiti: İlk {analiz_limiti} güçlü aday")

    print(f"\nKAP analizi başlıyor. İlk {analiz_limiti} güçlü aday kontrol edilecek...")
    kap_sonuclari = kap_toplu_analiz(zenginlestirilecek_liste, bekleme=0.15)

    print("\nÇoklu zaman dilimi analizi başlıyor...")
    mtf_sonuclari = {}
    for idx, sym in enumerate(zenginlestirilecek_liste, start=1):
        print(f"MTF analiz {idx}/{len(zenginlestirilecek_liste)}: {sym}")
        mtf_sonuclari[sym] = coklu_zaman_dilimi_analizi(sym)

    faaliyet_limiti = int(os.environ.get("FAALIYET_ANALIZ_LIMIT", "10"))
    faaliyet_liste = zenginlestirilecek_liste[:faaliyet_limiti]
    print(f"\nFaaliyet raporu analizi başlıyor. İlk {len(faaliyet_liste)} güçlü aday incelenecek...")
    faaliyet_sonuclari = faaliyet_toplu_analiz(faaliyet_liste, bekleme=0.35)

    yeni_results = []
    toplam = len(results)

    for i, item in enumerate(results, start=1):
        symbol = item["symbol"]

        if symbol in zenginlestirilecek:
            print(f"Pro analiz {i}/{toplam}: {symbol}")
            temel = temel_analiz_yfinance(symbol)
            haber = haber_analizi_yfinance(symbol)
            kap = kap_sonuclari.get(symbol, {
                "kap_skor": 0,
                "kap_etiket": "Veri Yok",
                "kap_notu": "KAP sonucu bulunamadı",
                "kap_basliklari": ""
            })
            mtf = mtf_sonuclari.get(symbol, {
                "mtf_skor": 50,
                "mtf_karar": "TUT",
                "mtf_uyum": "Veri Yok",
                "gunluk_yon": "",
                "gunluk_skor": 50,
                "gunluk_not": "",
                "haftalik_yon": "",
                "haftalik_skor": 50,
                "haftalik_not": ""
            })
            faaliyet = faaliyet_sonuclari.get(symbol, {
                "faaliyet_puani": 50,
                "faaliyet_gorunumu": "NÖTR",
                "faaliyet_raporu_bulundu": "Hayır",
                "faaliyet_raporu_basligi": "",
                "faaliyet_raporu_url": "",
                "faaliyet_raporu_dosya": "",
                "faaliyet_raporu_sayfa": 0,
                "faaliyet_raporu_karakter": 0,
                "faaliyet_metin_puani": 50,
                "faaliyet_guclu_noktalar": "",
                "faaliyet_riskleri": "",
                "faaliyet_notu": "Faaliyet raporu analizi ilk seçilen adaylarda uygulandı"
            })
        else:
            temel = {
                "temel_puan": 50, "fk": 0, "ileri_fk": 0, "pddd": 0,
                "borc_ozsermaye": 0, "roe": 0, "kar_marji": 0,
                "ciro_buyume": 0, "kar_buyume": 0, "temettu_verimi": 0,
                "temel_not": "Pro analiz ilk seçilen aday için uygulandı",
                "temel_risk": ""
            }
            haber = {
                "haber_puani": 0, "haber_etiket": "Nötr",
                "haber_notu": "Pro haber analizi ilk seçilen aday için uygulandı",
                "haber_basliklari": ""
            }
            kap = {
                "kap_skor": 0,
                "kap_etiket": "Nötr",
                "kap_notu": "KAP analizi seçilen adaylar için uygulandı",
                "kap_basliklari": ""
            }
            mtf = {
                "mtf_skor": 50,
                "mtf_karar": "TUT",
                "mtf_uyum": "Analiz dışı",
                "gunluk_yon": "",
                "gunluk_skor": 50,
                "gunluk_not": "",
                "haftalik_yon": "",
                "haftalik_skor": 50,
                "haftalik_not": ""
            }
            faaliyet = {
                "faaliyet_puani": 50,
                "faaliyet_gorunumu": "NÖTR",
                "faaliyet_raporu_bulundu": "Analiz dışı",
                "faaliyet_raporu_basligi": "",
                "faaliyet_raporu_url": "",
                "faaliyet_raporu_dosya": "",
                "faaliyet_raporu_sayfa": 0,
                "faaliyet_raporu_karakter": 0,
                "faaliyet_metin_puani": 50,
                "faaliyet_guclu_noktalar": "",
                "faaliyet_riskleri": "",
                "faaliyet_notu": "Faaliyet raporu analizi ilk 10 güçlü aday için uygulandı"
            }

        item.update(temel)
        item.update(haber)
        item.update(kap)
        item.update(mtf)
        item.update(faaliyet)
        item.update(makro)

        # KAP skorunu haber puanına da ek etki olarak yansıtıyoruz.
        item["haber_puani"] = item.get("haber_puani", 0) + item.get("kap_skor", 0)

        broker = broker_karar_motoru(item, makro)
        item.update(broker)

        # MTF uyumu broker skoruna küçük düzeltme uygular
        if item.get("mtf_uyum") == "Güçlü Uyum" and item.get("mtf_karar") == "AL":
            item["broker_skor"] = min(100, item.get("broker_skor", 0) + 5)
        elif item.get("mtf_uyum") == "Negatif Uyum":
            item["broker_skor"] = max(0, item.get("broker_skor", 0) - 8)

        # Faaliyet raporu puanı Broker skoruna kontrollü şekilde dahil edilir.
        faaliyet_puani = float(item.get("faaliyet_puani", 50))
        eski_broker = float(item.get("broker_skor", 50))
        item["broker_skor"] = int(round(max(0, min(100, eski_broker * 0.90 + faaliyet_puani * 0.10))))

        # Faaliyet görünümü çok güçlü veya çok zayıfsa aksiyonu kontrollü olarak düzelt.
        if item["broker_skor"] >= 78 and item.get("aksiyon") == "AL" and item.get("risk_getiri_1", 0) >= 1.5:
            item["broker_aksiyon"] = "GÜÇLÜ AL"
        elif item["broker_skor"] >= 68 and item.get("broker_aksiyon") != "SAT":
            item["broker_aksiyon"] = "AL"
        elif item["broker_skor"] <= 38:
            item["broker_aksiyon"] = "SAT"
        else:
            item["broker_aksiyon"] = "TUT"

        item["broker_yorum"] = (
            item.get("broker_yorum", "")
            + f" Faaliyet raporu puanı {faaliyet_puani:.0f}/100 ve görünüm "
            + f"{item.get('faaliyet_gorunumu', 'NÖTR')} olarak hesaba katıldı."
        )

        yeni_results.append(item)

    return yeni_results


def tabloya_cevir(results):
    tablo = []
    for item in results:
        tablo.append({
            "Hisse": item["symbol"],
            "v4 Görünüm": item.get("v4_gorunum", "NÖTR / İZLE"),
            "v4 Güven Puanı": item.get("v4_guven_puani", item.get("broker_skor", 0)),
            "v4 2-6 Hafta Puanı": item.get("v4_2_6_hafta_puani", 0),
            "v4 Nedenler": item.get("v4_nedenler", "Veri yok"),
            "v4 Uyarılar": item.get("v4_uyarilar", "Veri yok"),
            "Trend Katmanı": item.get("v4_trend_puani", 0),
            "Momentum Katmanı": item.get("v4_momentum_puani", 0),
            "Hacim Katmanı": item.get("v4_hacim_puani", 0),
            "Risk Katmanı": item.get("v4_risk_puani", 0),
            "Broker Aksiyon": item.get("broker_aksiyon", item.get("aksiyon", "TUT")),
            "Broker Skor": item.get("broker_skor", item.get("genel_skor", 0)),
            "Pozisyon Öneri %": item.get("pozisyon_orani_oneri", 0),
            "Aksiyon": item.get("aksiyon", "TUT"),
            "Güven": item.get("guven", 0),
            "Genel Skor": item.get("genel_skor", item.get("guven", 0)),
            "Puan": item.get("score", 0),
            "Karar": item.get("karar", ""),
            "Fiyat": round(item["price"], 2),
            "Alış Alt": round(item.get("alis_araligi_alt", 0), 2),
            "Alış Üst": round(item.get("alis_araligi_ust", 0), 2),
            "Stop Loss": round(item.get("stop_loss", 0), 2),
            "Hedef 1": round(item.get("hedef_1", 0), 2),
            "Hedef 2": round(item.get("hedef_2", 0), 2),
            "Risk/Getiri 1": round(item.get("risk_getiri_1", 0), 2),
            "Risk/Getiri 2": round(item.get("risk_getiri_2", 0), 2),
            "Tahmini Süre": item.get("tahmini_sure", ""),
            "Formasyon": item.get("formasyon", ""),
            "Formasyon Puanı": item.get("formasyon_puani", 0),
            "Formasyon Notu": item.get("formasyon_notu", ""),
            "Ana Destek": round(item.get("ana_destek", 0), 2),
            "Ana Direnç": round(item.get("ana_direnc", 0), 2),
            "Destek Gücü": item.get("destek_gucu", 0),
            "Direnç Gücü": item.get("direnc_gucu", 0),
            "Destek Mesafe %": round(item.get("destek_mesafe_yuzde", 0), 2),
            "Direnç Mesafe %": round(item.get("direnc_mesafe_yuzde", 0), 2),
            "Trend Yönü": item.get("trend_yonu", ""),
            "Trend Olasılığı": item.get("trend_olasiligi", 0),
            "15 Gün Tahmin": item.get("tahmin_15gun") or "Hesaplanamadı",
            "15 Gün Alt Bant": item.get("tahmin_alt", 0),
            "15 Gün Üst Bant": item.get("tahmin_ust", 0),
            "Teknik Risk Seviyesi": item.get("risk_seviyesi") or "Veri yok",
            "Algoritmik Teknik Yorum": item.get("ai_yorum") or "Teknik yorum üretilemedi",
            "Broker Yorum": item.get("broker_yorum", ""),
            "MTF Karar": item.get("mtf_karar", ""),
            "MTF Skor": item.get("mtf_skor", 50),
            "MTF Uyum": item.get("mtf_uyum", ""),
            "Günlük Yön": item.get("gunluk_yon", ""),
            "Günlük Skor": item.get("gunluk_skor", 50),
            "Haftalık Yön": item.get("haftalik_yon", ""),
            "Haftalık Skor": item.get("haftalik_skor", 50),
            "Grafik Dosyası": item.get("grafik_dosyasi", ""),
            "5 Gün %10+ Olasılık": item.get("olasilik_5g_10yukselis", 0),
            "10 Gün %10+ Olasılık": item.get("olasilik_10g_10yukselis", 0),
            "10 Gün Yeni Zirve Olasılık": item.get("olasilik_10g_yeni_zirve", 0),
            "Hedef 1 Olasılık": item.get("olasilik_hedef1", 0),
            "Stop Olasılık": item.get("olasilik_stop", 0),
            "10 Gün %20+ Olasılık": item.get("olasilik_10g_20yukselis", 0),
            "20 Gün %20+ Olasılık": item.get("olasilik_20g_20yukselis", 0),
            "30 Gün %20+ Olasılık": item.get("olasilik_30g_20yukselis", 0),
            "10 Gün %30+ Olasılık": item.get("olasilik_10g_30yukselis", 0),
            "20 Gün %30+ Olasılık": item.get("olasilik_20g_30yukselis", 0),
            "30 Gün %30+ Olasılık": item.get("olasilik_30g_30yukselis", 0),
            "Olasılık Örnek Sayısı": item.get("olasilik_ornek_sayisi", 0),
            "Olasılık Notu": item.get("olasilik_notu", ""),
            "Faaliyet Puanı": item.get("faaliyet_puani", 50),
            "Faaliyet Görünümü": item.get("faaliyet_gorunumu", "NÖTR"),
            "Faaliyet Raporu Bulundu": item.get("faaliyet_raporu_bulundu", "Hayır"),
            "Faaliyet Raporu Başlığı": item.get("faaliyet_raporu_basligi", ""),
            "Faaliyet Raporu Sayfa": item.get("faaliyet_raporu_sayfa", 0),
            "Faaliyet Metin Puanı": item.get("faaliyet_metin_puani", 50),
            "Faaliyet Güçlü Noktalar": item.get("faaliyet_guclu_noktalar", ""),
            "Faaliyet Riskleri": item.get("faaliyet_riskleri", ""),
            "Faaliyet Notu": item.get("faaliyet_notu", ""),
            "Faaliyet Raporu URL": item.get("faaliyet_raporu_url", ""),
            "Faaliyet PDF": item.get("faaliyet_raporu_dosya", ""),
            "Temel Puan": item.get("temel_puan", 50),
            "F/K": round(item.get("fk", 0), 2),
            "PD/DD": round(item.get("pddd", 0), 2),
            "Borç/Özsermaye": round(item.get("borc_ozsermaye", 0), 2),
            "ROE": round(item.get("roe", 0), 4),
            "Kâr Marjı": round(item.get("kar_marji", 0), 4),
            "Ciro Büyüme": round(item.get("ciro_buyume", 0), 4),
            "Kâr Büyüme": round(item.get("kar_buyume", 0), 4),
            "Temel Not": item.get("temel_not", ""),
            "Temel Risk": item.get("temel_risk", ""),
            "Haber Etiket": item.get("haber_etiket", ""),
            "Haber Puanı": item.get("haber_puani", 0),
            "Haber Notu": item.get("haber_notu", ""),
            "Haber Başlıkları": item.get("haber_basliklari", ""),
            "KAP Etiket": item.get("kap_etiket", ""),
            "KAP Skor": item.get("kap_skor", 0),
            "KAP Notu": item.get("kap_notu", ""),
            "KAP Başlıkları": item.get("kap_basliklari", ""),
            "Makro Etiket": item.get("makro_etiket", ""),
            "Makro Puan": item.get("makro_puan", 50),
            "Makro Notu": item.get("makro_notu", ""),
            "RSI": round(item.get("rsi", 0), 2),
            "ADX": round(item.get("adx", 0), 2),
            "ATR": round(item.get("atr", 0), 2),
            "EMA20": round(item.get("ema20", 0), 2),
            "EMA50": round(item.get("ema50", 0), 2),
            "EMA200": round(item.get("ema200", 0), 2),
            "MACD": round(item.get("macd", 0), 2),
            "MACD Signal": round(item.get("macd_signal", 0), 2),
            "Son 20 Gün %": round(item.get("ret_20", 0), 2),
            "Son 60 Gün %": round(item.get("ret_60", 0), 2),
            "Nedenler": " | ".join(item.get("reasons", [])),
            "Riskler": " | ".join(item.get("risk_notes", []))
        })
    df = pd.DataFrame(tablo)

    if not df.empty:
        # Ücretli yapay zekâ kullanmadan, mevcut analiz katmanlarından birleşik güven puanı.
        teknik = df["Güven"].fillna(50).clip(0, 100)
        temel = df["Temel Puan"].fillna(50).clip(0, 100)
        faaliyet = df["Faaliyet Puanı"].fillna(50).clip(0, 100) if "Faaliyet Puanı" in df.columns else 50
        mtf = df["MTF Skor"].fillna(50).clip(0, 100)
        makro = df["Makro Puan"].fillna(50).clip(0, 100)
        kap = (50 + df["KAP Skor"].fillna(0)).clip(0, 100)
        haber = (50 + df["Haber Puanı"].fillna(0)).clip(0, 100)
        rr = (df["Risk/Getiri 1"].fillna(0).clip(0, 3) / 3 * 100)

        df["AI Güven Puanı"] = (
            teknik * 0.30 + temel * 0.15 + faaliyet * 0.15 +
            mtf * 0.15 + kap * 0.08 + haber * 0.07 +
            makro * 0.05 + rr * 0.05
        ).round(1)

        fiyat = df["Fiyat"].replace(0, pd.NA)
        df["Beklenen Getiri %"] = (((df["Hedef 2"] / fiyat) - 1) * 100).fillna(0).round(1)

        def seviye(row):
            puan = float(row.get("AI Güven Puanı", 0))
            aksiyon = str(row.get("Broker Aksiyon", ""))
            if puan >= 82 and aksiyon in ["GÜÇLÜ AL", "AL"]:
                return "★★★★★ ÇOK GÜÇLÜ"
            if puan >= 74 and aksiyon in ["GÜÇLÜ AL", "AL"]:
                return "★★★★☆ GÜÇLÜ"
            if puan >= 64:
                return "★★★☆☆ İZLE"
            return "★★☆☆☆ ZAYIF"

        def neden(row):
            n = []
            if float(row.get("Güven", 0)) >= 70: n.append("Teknik güçlü")
            if float(row.get("MTF Skor", 0)) >= 65: n.append("Günlük/haftalık uyumlu")
            if float(row.get("Temel Puan", 0)) >= 65: n.append("Temel görünüm olumlu")
            if float(row.get("Faaliyet Puanı", 0)) >= 65: n.append("Faaliyet görünümü olumlu")
            if str(row.get("KAP Etiket", "")) == "Olumlu": n.append("KAP olumlu")
            if float(row.get("Risk/Getiri 1", 0)) >= 1.5: n.append("Risk/getiri uygun")
            if float(row.get("Beklenen Getiri %", 0)) >= 15: n.append("Hedef potansiyeli yüksek")
            return " | ".join(n[:5]) if n else "İzleme kriterleri sınırlı"

        df["Fırsat Seviyesi"] = df.apply(seviye, axis=1)
        df["Seçilme Nedenleri"] = df.apply(neden, axis=1)

    # Ekran ve raporda anlamsız boş hücre kalmaması için son kontrol.
    metin_kolonlari = df.select_dtypes(include=["object"]).columns
    for kolon in metin_kolonlari:
        df[kolon] = df[kolon].fillna("").astype(str).str.strip()
        df[kolon] = df[kolon].replace(
            {"": "Veri yok", "nan": "Veri yok", "None": "Veri yok"}
        )

    sayisal_kolonlar = df.select_dtypes(include=["number"]).columns
    df[sayisal_kolonlar] = df[sayisal_kolonlar].fillna(0)

    return df


def excel_stil_uygula(workbook):
    koyu = "111827"
    beyaz = "FFFFFF"
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False

        for cell in ws[1]:
            cell.font = Font(bold=True, color=beyaz)
            cell.fill = PatternFill("solid", fgColor=koyu)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            aksiyon = str(row[1].value) if len(row) > 1 else ""
            fill_color = None
            if aksiyon == "GÜÇLÜ AL":
                fill_color = "BBF7D0"
            elif aksiyon == "AL":
                fill_color = "DCFCE7"
            elif aksiyon == "TUT":
                fill_color = "FEF3C7"
            elif aksiyon == "SAT":
                fill_color = "FEE2E2"

            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if fill_color:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

        for col in ws.columns:
            max_len = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 55)

        ws.row_dimensions[1].height = 24


def guvenli_hucre_yaz(ws, hucre_adresi, deger):
    """
    Birleştirilmiş alan içindeki salt okunur hücreye yazmayı önler.
    Normal hücreye veya birleştirilmiş alanın sol üst hücresine yazar.
    """
    from openpyxl.cell.cell import MergedCell

    cell = ws[hucre_adresi]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if hucre_adresi in merged_range:
                sol_ust = ws.cell(merged_range.min_row, merged_range.min_col)
                sol_ust.value = deger
                return sol_ust
        raise ValueError(f"Yazılabilir hücre bulunamadı: {hucre_adresi}")

    cell.value = deger
    return cell


def dashboard_olustur(writer, df, baslangic_zamani, temettu_df=None):
    """
    v4.2 Dashboard:
    - Dekoratif boş hücrelere Veri yok yazılmaz.
    - KPI kartları ayrı bloklarda gösterilir.
    - İlk 10 aday sade sütunlarla listelenir.
    - Yaklaşan ilk 5 temettü ayrı bölümde gösterilir.
    """
    toplam = len(df)
    aksiyon_kolonu = "Broker Aksiyon" if "Broker Aksiyon" in df.columns else "Aksiyon"

    guclu_al = int((df[aksiyon_kolonu] == "GÜÇLÜ AL").sum()) if toplam and aksiyon_kolonu in df else 0
    al = int((df[aksiyon_kolonu] == "AL").sum()) if toplam and aksiyon_kolonu in df else 0
    tut = int((df[aksiyon_kolonu] == "TUT").sum()) if toplam and aksiyon_kolonu in df else 0
    sat = int((df[aksiyon_kolonu] == "SAT").sum()) if toplam and aksiyon_kolonu in df else 0
    ort_skor = round(float(df["v4 Güven Puanı"].mean()), 1) if toplam and "v4 Güven Puanı" in df else 0
    sure_dk = round((time.time() - baslangic_zamani) / 60, 1)

    ws = writer.book.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A11"

    # Ana başlık
    ws.merge_cells("A1:J2")
    ws["A1"] = "BORSA ANALİZ PRO MAX v4.2"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 12

    # KPI kart yardımcı fonksiyonu
    # Başlık ve değer alanları ayrı ayrı birleştirilir.
    # Böylece salt okunur MergedCell hücrelerine yazılmaz.
    def kart(baslik_araligi, deger_araligi, baslik, deger, renk):
        ws.merge_cells(baslik_araligi)
        ws.merge_cells(deger_araligi)

        baslik_hucre = baslik_araligi.split(":")[0]
        deger_hucre = deger_araligi.split(":")[0]

        for aralik in (baslik_araligi, deger_araligi):
            for row in ws[aralik]:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=renk)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws[baslik_hucre] = baslik
        ws[baslik_hucre].font = Font(size=10, bold=True, color="D1D5DB")
        ws[baslik_hucre].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        ws[deger_hucre] = deger
        ws[deger_hucre].font = Font(size=18, bold=True, color="FFFFFF")
        ws[deger_hucre].alignment = Alignment(horizontal="center", vertical="center")

    # Her kartın başlık ve değer bölümü ayrı birleştirilmiştir.
    kart("A4:B5", "A6:B7", "TOPLAM HİSSE", toplam, "1E3A8A")
    kart("C4:D5", "C6:D7", "GÜÇLÜ / POZİTİF", guclu_al + al, "166534")
    kart("E4:F5", "E6:F7", "NÖTR / İZLE", tut, "A16207")
    kart("G4:H5", "G6:H7", "NEGATİF / RİSKLİ", sat, "991B1B")
    kart("I4:J5", "I6:J7", "ORT. v4 GÜVEN", ort_skor, "6B21A8")

    ws.merge_cells("A9:J9")
    ws["A9"] = (
        f"Analiz süresi: {sure_dk} dk   |   "
        f"Rapor: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    ws["A9"].font = Font(italic=True, color="64748B")
    ws["A9"].alignment = Alignment(horizontal="center")

    # İlk 10 aday
    ws.merge_cells("A11:J11")
    ws["A11"] = "BUGÜNÜN EN YÜKSEK PUANLI 10 ADAYI"
    ws["A11"].font = Font(size=13, bold=True, color="FFFFFF")
    ws["A11"].fill = PatternFill("solid", fgColor="1E293B")
    ws["A11"].alignment = Alignment(horizontal="left")

    aday_kolonlari = [
        "Hisse", "v4 Görünüm", "v4 Güven Puanı", "v4 2-6 Hafta Puanı",
        "Fiyat", "Stop Loss", "Hedef 1", "Risk/Getiri 1",
        "MTF Karar", "v4 Nedenler"
    ]
    mevcut_aday_kolonlari = [c for c in aday_kolonlari if c in df.columns]
    ilk10 = df.head(10)[mevcut_aday_kolonlari].copy()
    ilk10.to_excel(writer, index=False, sheet_name="Dashboard", startrow=11)

    # Yaklaşan temettüler
    baslangic_satiri = 24
    ws.merge_cells(start_row=baslangic_satiri, start_column=1, end_row=baslangic_satiri, end_column=10)
    ws.cell(baslangic_satiri, 1, "EN YAKIN TEMETTÜLER")
    ws.cell(baslangic_satiri, 1).font = Font(size=13, bold=True, color="FFFFFF")
    ws.cell(baslangic_satiri, 1).fill = PatternFill("solid", fgColor="4C1D95")
    ws.cell(baslangic_satiri, 1).alignment = Alignment(horizontal="left")

    if temettu_df is not None and not temettu_df.empty:
        yaklasan = temettu_df.copy()
        if "Kalan Gün" in yaklasan.columns:
            yaklasan = yaklasan[
                pd.to_numeric(yaklasan["Kalan Gün"], errors="coerce").fillna(-1) >= 0
            ]
        temettu_kolonlari = [
            "Hisse", "Yaklaşan Temettü/Ex-Date", "Kalan Gün",
            "Temettü Verimi %", "Temettü Durumu"
        ]
        temettu_kolonlari = [c for c in temettu_kolonlari if c in yaklasan.columns]
        yaklasan.head(5)[temettu_kolonlari].to_excel(
            writer, index=False, sheet_name="Dashboard", startrow=baslangic_satiri
        )
    else:
        ws.cell(baslangic_satiri + 2, 1, "Yaklaşan temettü verisi bulunamadı.")
        ws.cell(baslangic_satiri + 2, 1).font = Font(color="64748B", italic=True)

    # Yasal uyarı
    uyari_satiri = 33
    ws.merge_cells(start_row=uyari_satiri, start_column=1, end_row=uyari_satiri + 1, end_column=10)
    ws.cell(uyari_satiri, 1, YASAL_UYARI_KISA)
    ws.cell(uyari_satiri, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(uyari_satiri, 1).fill = PatternFill("solid", fgColor="7F1D1D")
    ws.cell(uyari_satiri, 1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # Sütun genişlikleri, taşma olmadan kontrollü.
    genislikler = {
        "A": 14, "B": 22, "C": 15, "D": 18, "E": 12,
        "F": 12, "G": 12, "H": 14, "I": 13, "J": 42
    }
    for kolon, genislik in genislikler.items():
        ws.column_dimensions[kolon].width = genislik

    # Dashboard tablosu başlıkları
    for satir in [12, baslangic_satiri + 1]:
        for cell in ws[satir]:
            if cell.value is not None:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="334155")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Veri satırlarında hafif zebra görünümü.
    for row in range(13, 23):
        if row % 2 == 1:
            for col in range(1, 11):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="F8FAFC")

    # Tarih formatları
    for row in range(baslangic_satiri + 2, baslangic_satiri + 7):
        ws.cell(row, 2).number_format = "dd.mm.yyyy"

def potansiyel_adaylari_hazirla(df: pd.DataFrame):
    """
    Üç tablo üretir:
    1) Katı filtreyi geçen gerçek adaylar
    2) Katı filtreye yaklaşan izleme adayları
    3) Her koşulun sonucunu gösteren test/teşhis tablosu
    """
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    aday = df.copy()

    # Sıfıra bölme ve bozuk hedefleri önle
    fiyat = aday["Fiyat"].replace(0, pd.NA)
    aday["Hedef 1 Potansiyel %"] = (((aday["Hedef 1"] / fiyat) - 1) * 100).fillna(0).round(2)
    aday["Hedef 2 Potansiyel %"] = (((aday["Hedef 2"] / fiyat) - 1) * 100).fillna(0).round(2)

    rr = aday["Risk/Getiri 1"].fillna(0).clip(lower=0, upper=3)
    hist20 = aday["30 Gün %20+ Olasılık"].fillna(0).clip(lower=0, upper=100)
    hist30 = aday["30 Gün %30+ Olasılık"].fillna(0).clip(lower=0, upper=100)

    aday["2-6 Hafta Potansiyel Skor"] = (
        aday.get("v4 2-6 Hafta Puanı", aday["Broker Skor"]).fillna(0) * 0.35
        + aday["MTF Skor"].fillna(0) * 0.20
        + hist20 * 0.25
        + hist30 * 0.10
        + (rr / 3 * 100) * 0.10
    ).round(1)

    # Her şartı ayrı ayrı kaydet: test ekranında neden elendiği görülsün.
    aday["Şart 1 - AL"] = aday["Broker Aksiyon"].isin(["GÜÇLÜ AL", "AL"])
    aday["Şart 2 - v4 Güven ≥75"] = aday["v4 Güven Puanı"].fillna(0) >= 75
    aday["Şart 3 - MTF ≥65"] = aday["MTF Skor"].fillna(0) >= 65
    aday["Şart 4 - R/G ≥1.5"] = aday["Risk/Getiri 1"].fillna(0) >= 1.5
    aday["Şart 5 - Hedef2 ≥%20"] = aday["Hedef 2 Potansiyel %"] >= 20
    aday["Şart 6 - Tarihsel ≥%8"] = aday["30 Gün %20+ Olasılık"].fillna(0) >= 8

    sart_kolonlari = [
        "Şart 1 - AL", "Şart 2 - v4 Güven ≥75", "Şart 3 - MTF ≥65",
        "Şart 4 - R/G ≥1.5", "Şart 5 - Hedef2 ≥%20",
        "Şart 6 - Tarihsel ≥%8"
    ]

    aday["Geçen Şart Sayısı"] = aday[sart_kolonlari].sum(axis=1)

    def nedenler(row):
        mesajlar = []
        if not row["Şart 1 - AL"]:
            mesajlar.append("Aksiyon AL değil")
        if not row["Şart 2 - v4 Güven ≥75"]:
            mesajlar.append(f"v4 Güven {row['v4 Güven Puanı']:.0f}<75")
        if not row["Şart 3 - MTF ≥65"]:
            mesajlar.append(f"MTF {row['MTF Skor']:.0f}<65")
        if not row["Şart 4 - R/G ≥1.5"]:
            mesajlar.append(f"R/G {row['Risk/Getiri 1']:.2f}<1.5")
        if not row["Şart 5 - Hedef2 ≥%20"]:
            mesajlar.append(f"Hedef2 %{row['Hedef 2 Potansiyel %']:.1f}<%20")
        if not row["Şart 6 - Tarihsel ≥%8"]:
            mesajlar.append(f"Tarihsel %{row['30 Gün %20+ Olasılık']:.1f}<%8")
        return " | ".join(mesajlar) if mesajlar else "Tüm şartlar geçti"

    aday["Elenme Nedeni"] = aday.apply(nedenler, axis=1)

    kosul_20 = aday[sart_kolonlari].all(axis=1)
    kosul_30 = (
        kosul_20
        & (aday["Broker Skor"].fillna(0) >= 75)
        & (aday["Hedef 2 Potansiyel %"] >= 30)
        & (aday["30 Gün %30+ Olasılık"].fillna(0) >= 4)
    )

    aday["Potansiyel Sınıfı"] = "İZLE"
    aday.loc[kosul_20, "Potansiyel Sınıfı"] = "%20+ ADAY"
    aday.loc[kosul_30, "Potansiyel Sınıfı"] = "%30+ YÜKSEK RİSKLİ ADAY"

    kolonlar = [
        "Hisse", "Potansiyel Sınıfı", "2-6 Hafta Potansiyel Skor",
        "Geçen Şart Sayısı", "Elenme Nedeni",
        "Broker Aksiyon", "Broker Skor", "MTF Karar", "MTF Skor", "MTF Uyum",
        "Fiyat", "Alış Alt", "Alış Üst", "Stop Loss", "Hedef 1", "Hedef 2",
        "Hedef 1 Potansiyel %", "Hedef 2 Potansiyel %", "Risk/Getiri 1",
        "10 Gün %20+ Olasılık", "20 Gün %20+ Olasılık", "30 Gün %20+ Olasılık",
        "10 Gün %30+ Olasılık", "20 Gün %30+ Olasılık", "30 Gün %30+ Olasılık",
        "Olasılık Örnek Sayısı", "Olasılık Notu",
        "Formasyon", "Trend Yönü", "KAP Etiket", "KAP Skor",
        "Haber Etiket", "Temel Puan", "Faaliyet Puanı", "Broker Yorum"
    ] + sart_kolonlari

    mevcut = [c for c in kolonlar if c in aday.columns]

    kati = aday[kosul_20][mevcut].sort_values(
        ["2-6 Hafta Potansiyel Skor", "Broker Skor"],
        ascending=[False, False]
    )

    # 6 şarttan en az 4'ünü geçenler yakın adaydır.
    yakin = aday[(~kosul_20) & (aday["Geçen Şart Sayısı"] >= 4)][mevcut].sort_values(
        ["Geçen Şart Sayısı", "2-6 Hafta Potansiyel Skor", "Broker Skor"],
        ascending=[False, False, False]
    ).head(30)

    # Test sayfasında olasılık analizi yapılan ilk 30 hisseyi öne al.
    test = aday[mevcut].sort_values(
        ["Geçen Şart Sayısı", "2-6 Hafta Potansiyel Skor", "Broker Skor"],
        ascending=[False, False, False]
    ).head(60)

    return kati, yakin, test


def bugunun_firsatlari_hazirla(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    aday = df.copy()
    aday = aday[aday["Broker Aksiyon"].isin(["GÜÇLÜ AL", "AL", "TUT"])].copy()
    aday = aday[aday["AI Güven Puanı"] >= 58].copy()

    kolonlar = [
        "Hisse", "Fırsat Seviyesi", "AI Güven Puanı", "Broker Aksiyon",
        "Broker Skor", "Fiyat", "Alış Alt", "Alış Üst", "Stop Loss",
        "Hedef 1", "Hedef 2", "Beklenen Getiri %", "Tahmini Süre",
        "Risk/Getiri 1", "MTF Karar", "MTF Skor", "Temel Puan",
        "Faaliyet Puanı", "Faaliyet Görünümü", "KAP Etiket", "Haber Etiket",
        "Seçilme Nedenleri", "Broker Yorum"
    ]
    mevcut = [c for c in kolonlar if c in aday.columns]
    return aday.sort_values(
        ["AI Güven Puanı", "Broker Skor", "Risk/Getiri 1"],
        ascending=[False, False, False]
    )[mevcut].head(10)

def sonuclari_kaydet(results, baslangic_zamani, backtest_ozet=None, backtest_islemler=None, temettu_df=None):
    results = sonuclari_sirala(results)
    df = tabloya_cevir(results)
    potansiyel_df, yakin_adaylar_df, potansiyel_test_df = potansiyel_adaylari_hazirla(df)
    firsatlar_df = bugunun_firsatlari_hazirla(df)

    output_dir = output_klasoru()

    excel_path = output_dir / "Borsa_Analiz_Pro_MAX_Rapor.xlsx"
    txt_path = output_dir / "Borsa_Analiz_Pro_MAX_Rapor.txt"

    portfoy = portfoy_onerisi_uret(df, sermaye=100000)
    faaliyet_df = faaliyet_dataframe(results)

    excel_path_fallback = None
    try:
        writer_context = pd.ExcelWriter(excel_path, engine="openpyxl")
    except PermissionError:
        zaman = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path_fallback = output_dir / f"Borsa_Analiz_Pro_MAX_Rapor_{zaman}.xlsx"
        writer_context = pd.ExcelWriter(excel_path_fallback, engine="openpyxl")

    with writer_context as writer:
        pd.DataFrame(YASAL_UYARI_UZUN, columns=["UYARI"]).to_excel(writer, index=False, sheet_name="Yasal Uyari")
        dashboard_olustur(writer, df, baslangic_zamani, temettu_df)
        firsatlar_df.to_excel(writer, index=False, sheet_name="Bugunun Firsatlari")
        df.to_excel(writer, index=False, sheet_name="Tum Sonuclar")
        df[df["Broker Aksiyon"] == "GÜÇLÜ AL"].head(50).to_excel(writer, index=False, sheet_name="Guclu AL")
        df[df["Broker Aksiyon"] == "AL"].head(50).to_excel(writer, index=False, sheet_name="AL")
        df[df["Broker Aksiyon"] == "TUT"].head(50).to_excel(writer, index=False, sheet_name="TUT")
        df[df["Broker Aksiyon"] == "SAT"].head(50).to_excel(writer, index=False, sheet_name="SAT")
        if not portfoy.empty:
            portfoy.to_excel(writer, index=False, sheet_name="Portfoy Onerisi")

        if backtest_ozet is not None and not backtest_ozet.empty:
            backtest_ozet.to_excel(writer, index=False, sheet_name="Backtest Ozet")

        if backtest_islemler is not None and not backtest_islemler.empty:
            backtest_islemler.to_excel(writer, index=False, sheet_name="Backtest Islemler")

        if temettu_df is not None and not temettu_df.empty:
            temettu_df.to_excel(writer, index=False, sheet_name="Temettu Takip")

        # Katı filtre sonucu boş olsa bile sayfa oluşturulur.
        potansiyel_df.to_excel(writer, index=False, sheet_name="2-6 Hafta Potansiyel")
        yakin_adaylar_df.to_excel(writer, index=False, sheet_name="2-6 Hafta Yakin")

        if not faaliyet_df.empty:
            faaliyet_df.to_excel(writer, index=False, sheet_name="Faaliyet Raporlari")

        # Bütün Excel sayfalarında boş hücreleri anlaşılır metinle göster.
        # Birleştirilmiş hücrelerin MergedCell nesneleri salt okunurdur.
        from openpyxl.cell.cell import MergedCell

        for ws in writer.book.worksheets:
            # Dashboard'daki boşluklar tasarım amaçlıdır; "Veri yok" ile doldurulmaz.
            if ws.title == "Dashboard":
                continue

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    # Yalnızca gerçek veri tablosunun içinde kalan boş hücreleri doldur.
                    if cell.value is None and cell.row >= 2 and cell.column <= ws.max_column:
                        cell.value = "Veri yok"

        excel_stil_uygula(writer.book)

        # Temettü tarih ve kalan gün kolonlarını okunaklı biçimlendir.
        if "Temettu Takip" in writer.book.sheetnames:
            tws = writer.book["Temettu Takip"]
            basliklar = {cell.value: cell.column for cell in tws[1]}
            tarih_col = basliklar.get("Yaklaşan Temettü/Ex-Date")
            kalan_col = basliklar.get("Kalan Gün")

            if tarih_col:
                for row in range(2, tws.max_row + 1):
                    tws.cell(row, tarih_col).number_format = "dd.mm.yyyy"

            if kalan_col:
                for row in range(2, tws.max_row + 1):
                    cell = tws.cell(row, kalan_col)
                    try:
                        kalan = int(cell.value)
                        if kalan == 0:
                            cell.fill = PatternFill("solid", fgColor="FDE68A")
                        elif 0 < kalan <= 7:
                            cell.fill = PatternFill("solid", fgColor="DCFCE7")
                        elif kalan < 0:
                            cell.fill = PatternFill("solid", fgColor="F1F5F9")
                    except (TypeError, ValueError):
                        pass

    with txt_path.open("w", encoding="utf-8-sig") as f:
        f.write(YASAL_UYARI_KISA + "\n\n")
        df.to_csv(f, index=False, sep=";")

    print("\nDosyalar oluşturuldu:")
    print(f"- {excel_path_fallback or excel_path}")
    print(f"- {txt_path}")


def ozet_yazdir(results, baslangic_zamani):
    sirali = sonuclari_sirala(results)
    sure = time.time() - baslangic_zamani

    print("\n==============================")
    print(" PRO MAX ÖZET")
    print("==============================")
    print(f"Başarılı analiz: {len(results)}")
    print(f"Süre: {sure/60:.1f} dakika")

    for i, item in enumerate(sirali[:10], start=1):
        print(
            f"{i}. {item['symbol']} | {item.get('broker_aksiyon')} | "
            f"Broker {item.get('broker_skor')}/100 | "
            f"Fiyat {item.get('price', 0):.2f} | Stop {item.get('stop_loss', 0):.2f} | "
            f"Hedef1 {item.get('hedef_1', 0):.2f}"
        )


def main():
    baslangic_zamani = time.time()
    print("Borsa Analiz Pro MAX v4.2.1 başladı:", datetime.now().strftime("%d.%m.%Y %H:%M"))

    hisseler = hisseleri_txt_oku()
    print(f"Toplam taranacak hisse: {len(hisseler)}")

    results = []
    max_worker = 6

    with ThreadPoolExecutor(max_workers=max_worker) as executor:
        islemler = {executor.submit(hisse_tara, symbol): symbol for symbol in hisseler}
        tamamlanan = 0

        for future in as_completed(islemler):
            symbol = islemler[future]
            tamamlanan += 1
            try:
                sonuc = future.result()
                if sonuc:
                    results.append(sonuc)
                    print(f"{tamamlanan}/{len(hisseler)} teknik tamamlandı: {symbol} | {sonuc.get('aksiyon')} | Güven {sonuc.get('guven')}")
                else:
                    print(f"{tamamlanan}/{len(hisseler)} atlandı: {symbol}")
            except Exception as e:
                print(f"{tamamlanan}/{len(hisseler)} hata: {symbol} -> {e}")

    dogrulanmis_listeyi_kaydet(results)
    results = profesyonel_veri_ekle(results)
    results = v4_toplu_puanla(results, final=False)
    results = sonuclari_sirala(results)

    rapor_yazdir(results, limit=30)

    print("\nBacktest başlıyor. İlk 40 güçlü aday test edilecek...")
    backtest_hisseler = [x["symbol"] for x in results[:40]]
    backtest_ozet, backtest_islemler = backtest_toplu(backtest_hisseler, max_hisse=40)

    print("\nGrafik üretimi başlıyor. İlk 20 güçlü aday için PNG grafik oluşturulacak...")
    grafikler = grafik_toplu_olustur(results, limit=20, output_dir=str(output_klasoru() / "grafikler"))
    for item in results:
        sym = item.get("symbol")
        if sym in grafikler:
            item["grafik_dosyasi"] = grafikler[sym]

    print("\nTarihsel olasılık analizi başlıyor. İlk 30 güçlü aday hesaplanacak...")
    results = olasilik_toplu_ekle(results, limit=30)
    results = v4_toplu_puanla(results, final=True)
    results = sonuclari_sirala(results)

    print("\nTemettü taraması başlıyor...")
    temettu_df = temettu_toplu_tara(hisseleri_txt_oku(), max_workers=8)

    try:
        sonuclari_kaydet(
            results,
            baslangic_zamani,
            backtest_ozet,
            backtest_islemler,
            temettu_df
        )
    except Exception as exc:
        print(f"Excel raporu oluşturulamadı: {exc}")
        print("Sonuçlar acil CSV yedeğine kaydediliyor...")
        try:
            acil_df = tabloya_cevir(results)
            acil_yol = output_klasoru() / (
                "Borsa_Analiz_Pro_MAX_ACIL_YEDEK_"
                + datetime.now().strftime("%Y%m%d_%H%M%S")
                + ".csv"
            )
            acil_df.to_csv(acil_yol, index=False, encoding="utf-8-sig")
            print(f"Acil yedek kaydedildi: {acil_yol}")
        except Exception as yedek_exc:
            print(f"Acil yedek de oluşturulamadı: {yedek_exc}")
        raise

    ozet_yazdir(results, baslangic_zamani)
    print("\nPro MAX tarama tamamen bitti.")


if __name__ == "__main__":
    main()
